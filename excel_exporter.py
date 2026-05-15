"""
Excel Diagnostic Workbook Exporter
=====================================
Version: v1.1

Exports DuckDB tables to a multi-sheet Excel workbook for data validation.
Open this file after each pipeline run to inspect the data.

Sheets produced:
  1. Mapped Sales       — most recent 200,000 in-scope residential transactions
  2. Annual Metrics     — year-by-year suburb aggregations (small, always fits)
  3. Price Performance  — period comparisons for all 92 suburbs (small, always fits)
  4. Regional Metrics   — year-by-year region aggregations (small, always fits)
  5. Excluded Records   — sample of records filtered out (capped at 50,000)
  6. Ingestion Log      — which DAT files have been loaded and when

Note on row limits:
  Excel maximum is 1,048,576 rows per sheet.
  Mapped Sales and Excluded Records are sampled to stay well within this limit.
  All aggregated sheets (Annual Metrics, Price Performance, Regional Metrics)
  are naturally small and will never approach the limit.
"""

import logging
from datetime import datetime
from pathlib import Path

logger = logging.getLogger(__name__)


def export_excel(db_path: str, output_path: str) -> dict:
    """
    Export diagnostic tables to an Excel workbook.

    Args:
        db_path:     Path to DuckDB database
        output_path: Full path for the output .xlsx file

    Returns:
        Summary dict with row counts per sheet.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.error("openpyxl not installed. Run: pip install openpyxl")
        return {"error": "openpyxl not installed"}

    try:
        import duckdb
    except ImportError:
        logger.error("duckdb not installed. Run: pip install duckdb")
        return {"error": "duckdb not installed"}

    import time
    start = time.time()

    conn = duckdb.connect(db_path, read_only=True)
    wb   = openpyxl.Workbook()
    wb.remove(wb.active)

    summary = {}

    HEADER_FILL  = PatternFill("solid", fgColor="1F3864")
    HEADER_FONT  = Font(color="FFFFFF", bold=True, name="Calibri", size=10)
    CELL_FONT    = Font(name="Calibri", size=10)
    ALT_ROW_FILL = PatternFill("solid", fgColor="F2F2F2")

    def add_sheet(title: str, headers: list, rows: list) -> int:
        ws = wb.create_sheet(title=title)

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for row_idx, row in enumerate(rows, 2):
            fill = ALT_ROW_FILL if row_idx % 2 == 0 else None
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = CELL_FONT
                if fill:
                    cell.fill = fill

        for col_idx, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_idx)
            max_len = len(header)
            for row_data in rows[:200]:
                val = str(row_data[col_idx - 1]) if row_data[col_idx - 1] is not None else ""
                max_len = max(max_len, len(val))
            ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 18
        return len(rows)

    # ── Sheet 1: Mapped Sales (most recent 200,000 records) ───────
    try:
        total_mapped = conn.execute("SELECT COUNT(*) FROM mapped_sales").fetchone()[0]
        rows = conn.execute("""
            SELECT
                ms.suburb,
                ms.postcode,
                ms.region,
                ms.property_type,
                ms.contract_date,
                ms.sale_price,
                ms.land_size_sqm,
                CASE WHEN ms.land_size_sqm > 0
                     THEN ROUND(ms.sale_price / ms.land_size_sqm, 0)
                     ELSE NULL END AS price_per_sqm,
                ms.source_file,
                ms.contract_year
            FROM mapped_sales ms
            ORDER BY ms.contract_date DESC
            LIMIT 200000
        """).fetchall()

        headers = [
            "Suburb", "Postcode", "Region", "Property Type",
            "Contract Date", "Sale Price ($)", "Land Size (sqm)",
            "Price per sqm ($)", "Source File", "Year"
        ]
        n = add_sheet("Mapped Sales", headers, rows)
        summary["mapped_sales"] = n
        if total_mapped > 200000:
            logger.info("Sheet 'Mapped Sales': showing %d most recent of %d total records",
                        n, total_mapped)
        else:
            logger.info("Sheet 'Mapped Sales': %d rows", n)
    except Exception as exc:
        logger.warning("Could not export Mapped Sales: %s", exc)
        summary["mapped_sales"] = 0

    # ── Sheet 2: Annual Metrics ───────────────────────────────────
    try:
        rows = conn.execute("""
            SELECT
                suburb, postcode, region, contract_year,
                sales_count,
                ROUND(median_price, 0) AS median_price,
                ROUND(avg_price, 0)    AS avg_price,
                min_price, max_price,
                ROUND(median_land_sqm, 1) AS median_land_sqm
            FROM annual_suburb_metrics
            ORDER BY suburb, contract_year
        """).fetchall()

        headers = [
            "Suburb", "Postcode", "Region", "Year",
            "Sales Count", "Median Price ($)", "Avg Price ($)",
            "Min Price ($)", "Max Price ($)", "Median Land (sqm)"
        ]
        n = add_sheet("Annual Metrics", headers, rows)
        summary["annual_metrics"] = n
        logger.info("Sheet 'Annual Metrics': %d rows", n)
    except Exception as exc:
        logger.warning("Could not export Annual Metrics: %s", exc)
        summary["annual_metrics"] = 0

    # ── Sheet 3: Price Performance ────────────────────────────────
    try:
        rows = conn.execute("""
            SELECT
                suburb, postcode, region,
                sales_3m,  ROUND(median_3m, 0),  ROUND(pct_change_3m, 2),
                sales_12m, ROUND(median_12m, 0), ROUND(pct_change_12m, 2),
                sales_3y,  ROUND(median_3y, 0),  ROUND(pct_change_3y, 2),
                sales_5y,  ROUND(median_5y, 0),  ROUND(pct_change_5y, 2),
                sales_10y, ROUND(median_10y, 0), ROUND(pct_change_10y, 2),
                sales_20y, ROUND(median_20y, 0), ROUND(pct_change_20y, 2),
                as_of_date
            FROM price_performance
            ORDER BY suburb
        """).fetchall()

        headers = [
            "Suburb", "Postcode", "Region",
            "3m Sales", "3m Median ($)", "3m Change (%)",
            "12m Sales", "12m Median ($)", "12m Change (%)",
            "3y Sales", "3y Median ($)", "3y Change (%)",
            "5y Sales", "5y Median ($)", "5y Change (%)",
            "10y Sales", "10y Median ($)", "10y Change (%)",
            "20y Sales", "20y Median ($)", "20y Change (%)",
            "As Of Date"
        ]
        n = add_sheet("Price Performance", headers, rows)
        summary["price_performance"] = n
        logger.info("Sheet 'Price Performance': %d rows", n)
    except Exception as exc:
        logger.warning("Could not export Price Performance: %s", exc)
        summary["price_performance"] = 0

    # ── Sheet 4: Regional Metrics ─────────────────────────────────
    try:
        rows = conn.execute("""
            SELECT
                region, contract_year, sales_count,
                ROUND(median_price, 0) AS median_price,
                ROUND(avg_price, 0)    AS avg_price,
                suburb_count
            FROM annual_regional_metrics
            ORDER BY region, contract_year
        """).fetchall()

        headers = [
            "Region", "Year", "Sales Count",
            "Median Price ($)", "Avg Price ($)", "Active Suburbs"
        ]
        n = add_sheet("Regional Metrics", headers, rows)
        summary["regional_metrics"] = n
        logger.info("Sheet 'Regional Metrics': %d rows", n)
    except Exception as exc:
        logger.warning("Could not export Regional Metrics: %s", exc)
        summary["regional_metrics"] = 0

    # ── Sheet 5: Excluded Records (sample, capped at 50,000) ──────
    try:
        total_raw = conn.execute("SELECT COUNT(*) FROM raw_sales").fetchone()[0]
        rows = conn.execute("""
            SELECT
                rs.suburb,
                rs.postcode,
                rs.property_category,
                rs.property_description,
                rs.property_type,
                rs.sale_price,
                rs.contract_date,
                rs.source_file,
                CASE
                    WHEN rs.property_category != 'R'
                        THEN 'Not residential'
                    WHEN rs.sale_price IS NULL OR rs.sale_price <= 0
                        THEN 'No valid sale price'
                    ELSE 'Suburb not in approved universe'
                END AS exclusion_reason
            FROM raw_sales rs
            LEFT JOIN mapped_sales ms ON rs.id = ms.id
            WHERE ms.id IS NULL
            ORDER BY rs.contract_date DESC
            LIMIT 50000
        """).fetchall()

        headers = [
            "Suburb", "Postcode", "Property Category",
            "Property Description", "Property Type",
            "Sale Price", "Contract Date", "Source File",
            "Exclusion Reason"
        ]
        n = add_sheet("Excluded Records", headers, rows)
        summary["excluded_records"] = n
        logger.info("Sheet 'Excluded Records': %d rows (sample from %d total raw records)",
                    n, total_raw)
    except Exception as exc:
        logger.warning("Could not export Excluded Records: %s", exc)
        summary["excluded_records"] = 0

    # ── Sheet 6: Ingestion Log ────────────────────────────────────
    try:
        rows = conn.execute("""
            SELECT
                source_file,
                district_code,
                file_release_date,
                records_loaded,
                ingested_at
            FROM ingested_files
            ORDER BY ingested_at DESC
        """).fetchall()

        headers = [
            "Source File", "District Code", "File Release Date",
            "Records Loaded", "Ingested At"
        ]
        n = add_sheet("Ingestion Log", headers, rows)
        summary["ingestion_log"] = n
        logger.info("Sheet 'Ingestion Log': %d rows", n)
    except Exception as exc:
        logger.warning("Could not export Ingestion Log: %s", exc)
        summary["ingestion_log"] = 0

    conn.close()

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)

    elapsed = time.time() - start
    summary["elapsed_seconds"] = elapsed
    summary["output_path"]     = output_path

    logger.info("Excel workbook saved to %s (%.1fs)", output_path, elapsed)
    return summary
